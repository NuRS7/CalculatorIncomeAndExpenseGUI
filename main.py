import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import numpy as np
from mpl_toolkits.mplot3d import Axes3D
import os
import pandas as pd
import openpyxl
from openpyxl.chart import (
    PieChart, BarChart, Reference,
    PieChart3D, BarChart3D
)
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from tkcalendar import DateEntry


class FinanceTracker:
    def __init__(self, root):
        self.root = root
        self.root.title("💰 Kişisel Finans Yönetimi Pro")
        
        # Make window responsive
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        window_width = int(screen_width * 0.85)
        window_height = int(screen_height * 0.85)
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # Allow window resizing
        self.root.minsize(1000, 700)
        
        # Configure root grid for responsive layout
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        # Initialize database
        self.db = Database()
        
        # Color scheme
        self.colors = {
            'primary': '#2c3e50',
            'secondary': '#3498db',
            'success': '#2ecc71',
            'danger': '#e74c3c',
            'warning': '#f39c12',
            'light': '#ecf0f1',
            'dark': '#34495e',
            'bg': '#ffffff'
        }
        
        # Configure styles
        self.setup_styles()

        # Create main container with responsive grid
        self.main_frame = ttk.Frame(root, style='Main.TFrame')
        self.main_frame.grid(row=0, column=0, sticky='nsew', padx=10, pady=10)
        self.main_frame.grid_rowconfigure(0, weight=1)
        self.main_frame.grid_columnconfigure(0, weight=1)

        # Create menu
        self.create_menu()
        # Show dashboard by default
        self.show_dashboard()
    
    def export_to_excel(self):
        try:
            # Create a Pandas Excel writer
            file_path = os.path.join(os.path.expanduser('~'), 'Downloads', 'finance_report.xlsx')
            writer = pd.ExcelWriter(file_path, engine='openpyxl')
            
            # Get all transactions
            transactions = self.db.get_all_transactions()
            if not transactions:
                messagebox.showinfo(" Data yok", "Dışa aktarılacak işlem bulunamadı.")
                return
                
            # Convert to DataFrame
            df = pd.DataFrame(transactions)
            
            # Write transactions to the first sheet
            df.to_excel(writer, sheet_name='İşlemler', index=False)
            
            # Get category summary
            category_summary = self.db.get_category_summary()
            if category_summary:
                df_summary = pd.DataFrame(category_summary)
                df_summary.to_excel(writer, sheet_name='Kategori Özeti', index=False)
            
            # Get workbook and worksheets
            workbook = writer.book
            ws_trans = writer.sheets['İşlemler']
            
            # Format the transactions sheet
            for column in ws_trans.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws_trans.column_dimensions[column_letter].width = min(adjusted_width, 30)
            
            # Add summary statistics
            if 'Kategori Özeti' in writer.sheets:
                ws_summary = writer.sheets['Kategori Özeti']
                
                # Create charts
                self._create_pie_chart(workbook, ws_summary, df_summary, 'Gelir', 'Gelir Tablosu')
                self._create_pie_chart(workbook, ws_summary, df_summary, 'Gider', 'Gider Tablosu')
                
                # Add summary statistics
                total_income = df_summary[df_summary['type'] == 'Gelir']['total_amount'].sum()
                total_expenses = df_summary[df_summary['type'] == 'Gider']['total_amount'].sum()
                net_balance = total_income - total_expenses
                
                stats = [
                    ['Toplam Gelir', total_income],
                    ['Toplam Gider', total_expenses],
                    ['Net Bakiye', net_balance],
                    ['Tasarruf Oranı', f"{(net_balance/total_income*100 if total_income > 0 else 0):.1f}%" if total_income > 0 else 'N/A']
                ]
                
                # Add statistics to the summary sheet
                ws_summary['E1'] = 'İstatistikler'
                ws_summary['E1'].font = Font(bold=True, size=12)
                
                for i, (label, value) in enumerate(stats, start=2):
                    ws_summary[f'E{i}'] = label
                    ws_summary[f'F{i}'] = value
                    ws_summary[f'E{i}'].font = Font(bold=True)
                    if i == 3:  # Net Balance row
                        ws_summary[f'F{i}'].font = Font(bold=True, color='FF0000' if net_balance < 0 else '008000')
            
            # Save the Excel file
            writer.close()
            
            messagebox.showinfo("Dışa aktarma başarılı",
                              f"Rapor şu konuma başarıyla dışa aktarıldı:\n{file_path}")
            
        except Exception as e:
            messagebox.showerror("Dışa aktarma Hatası", f"Excel'e dışa aktarılırken bir hata oluştu:\n{str(e)}")
    
    def _create_pie_chart(self, workbook, ws, df, trans_type, title):
        # Filter data for the specified transaction type
        df_filtered = df[df['type'] == trans_type]
        if df_filtered.empty:
            return
            
        # Create a pie chart
        pie = PieChart3D()
        labels = Reference(ws, min_col=1, min_row=2, max_row=len(df_filtered)+1)
        data = Reference(ws, min_col=3, min_row=1, max_row=len(df_filtered)+1)
        
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = title
        
        # Position the chart
        max_row = ws.max_row + 2
        ws.add_chart(pie, f"E{max_row}")
        
        return pie

    def setup_styles(self):
        style = ttk.Style()
        style.theme_use('clam')
        
        # Main frame style
        style.configure('Main.TFrame', background=self.colors['light'])
        
        # Custom button styles
        style.configure('Primary.TButton',
                      background=self.colors['primary'],
                      foreground='white',
                      font=('Segoe UI', 11, 'bold'),
                      borderwidth=0,
                      focuscolor='none',
                      padding=10)
        
        style.configure('Success.TButton',
                      background=self.colors['success'],
                      foreground='white',
                      font=('Segoe UI', 11, 'bold'),
                      borderwidth=0,
                      padding=10)
        
        style.configure('Danger.TButton',
                      background=self.colors['danger'],
                      foreground='white',
                      font=('Segoe UI', 11, 'bold'),
                      borderwidth=0,
                      padding=10)

    def create_menu(self):
        menubar = tk.Menu(self.root)

        # File menu
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Excel'e Aktar", command=self.export_to_excel)
        file_menu.add_command(label="Pano", command=self.show_dashboard)
        file_menu.add_separator()
        file_menu.add_command(label="Çıkış", command=self.root.quit)

        # Transaction menu
        trans_menu = tk.Menu(menubar, tearoff=0)
        trans_menu.add_command(label="Gelir Ekle", command=lambda: self.show_add_transaction("Income"))
        trans_menu.add_command(label="Gider Ekle", command=lambda: self.show_add_transaction("Expense"))
        trans_menu.add_separator()
        trans_menu.add_command(label="Tüm İşlemleri Görüntüle", command=self.show_transactions)

        # Reports menu
        report_menu = tk.Menu(menubar, tearoff=0)
        report_menu.add_command(label="Günlük Rapor", command=lambda: self.show_report("Daily"))
        report_menu.add_command(label="Aylık Rapor", command=lambda: self.show_report("Monthly"))
        report_menu.add_command(label="Yıllık Rapor", command=lambda: self.show_report("Yearly"))

        # Analytics menu
        analytics_menu = tk.Menu(menubar, tearoff=0)
        analytics_menu.add_command(label="İstatistiksel Analiz", command=self.show_statistics)
        analytics_menu.add_command(label="3 Boyutlu Görselleştirme", command=self.show_3d_analysis)

        # Categories menu
        cat_menu = tk.Menu(menubar, tearoff=0)
        cat_menu.add_command(label="Kategorileri Yönet", command=self.manage_categories)

        # Add menus to menubar
        menubar.add_cascade(label="Dosya", menu=file_menu)
        menubar.add_cascade(label="İşlemler", menu=trans_menu)
        menubar.add_cascade(label="Raporlar", menu=report_menu)
        menubar.add_cascade(label="Analitik", menu=analytics_menu)
        menubar.add_cascade(label="Analitik", menu=cat_menu)

        self.root.config(menu=menubar)

    def clear_frame(self):
        for widget in self.main_frame.winfo_children():
            widget.destroy()
    


    def show_dashboard(self):
        self.clear_frame()
        
        # Create main container with PanedWindow for resizable sections
        main_paned = ttk.PanedWindow(self.main_frame, orient=tk.VERTICAL)
        main_paned.pack(fill=tk.BOTH, expand=True)
        
        # Top section - Stats and info
        top_frame = tk.Frame(main_paned, bg=self.colors['light'])
        main_paned.add(top_frame, weight=1)
        
        # Dashboard header
        header_frame = tk.Frame(top_frame, bg=self.colors['primary'], height=80)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        tk.Label(header_frame, text="📊 Pano",
                font=('Segoe UI', 28, 'bold'),
                bg=self.colors['primary'],
                fg='white').pack(side=tk.LEFT, padx=30, pady=20)
        
        # Balance display
        balance = self.db.get_balance()
        balance_color = self.colors['success'] if balance >= 0 else self.colors['danger']
        
        tk.Label(header_frame, text=f"💰 Balance: ${balance:,.2f}",
                font=('Segoe UI', 20, 'bold'),
                bg=self.colors['primary'],
                fg=balance_color).pack(side=tk.RIGHT, padx=30, pady=20)
        
        # Stats cards in scrollable frame
        stats_container = tk.Frame(top_frame, bg=self.colors['light'])
        stats_container.pack(fill=tk.X, padx=20, pady=15)
        
        # Get statistics
        total_income = self.db.get_total_income()
        total_expenses = self.db.get_total_expenses()
        monthly_income = self.db.get_monthly_income()
        monthly_expenses = self.db.get_monthly_expenses()
        
        stats_data = [
            ("📈", "Toplam Gelir", f"${total_income:,.2f}", self.colors['success']),
            ("📉", "Toplam Gider", f"${total_expenses:,.2f}", self.colors['danger']),
            ("📅", "Aylık Gelir", f"${monthly_income:,.2f}", self.colors['secondary']),
            ("💸", "Aylık Gider", f"${monthly_expenses:,.2f}", self.colors['warning'])
        ]
        
        for i, (icon, title, value, color) in enumerate(stats_data):
            card = tk.Frame(stats_container, bg='white', relief='solid', bd=1)
            card.grid(row=0, column=i, padx=10, pady=10, sticky='nsew')
            stats_container.grid_columnconfigure(i, weight=1)
            
            tk.Label(card, text=icon, font=('Segoe UI', 32),
                    bg='white').pack(pady=(15, 5))
            tk.Label(card, text=title, font=('Segoe UI', 10),
                    bg='white', fg='gray').pack()
            tk.Label(card, text=value, font=('Segoe UI', 18, 'bold'),
                    bg='white', fg=color).pack(pady=(5, 15))
        
        # Bottom section - Charts (resizable)
        bottom_frame = tk.Frame(main_paned, bg=self.colors['light'])
        main_paned.add(bottom_frame, weight=3)
        
        # Charts with PanedWindow for horizontal resizing
        charts_paned = ttk.PanedWindow(bottom_frame, orient=tk.HORIZONTAL)
        charts_paned.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Add charts
        self.create_resizable_charts(charts_paned)
    
    def create_resizable_charts(self, parent_paned):
        """Create resizable charts with interactive tooltips"""
        
        # Get data
        expense_data = self.db.get_expenses_by_category()
        monthly_data = self.db.get_monthly_summary()
        
        if not monthly_data or len(monthly_data) == 0:
            no_data_label = tk.Label(parent_paned, text="Veri yok. Grafik görmek için işlem ekleyin!",
                                    font=('Segoe UI', 14), bg=self.colors['light'], fg='gray')
            parent_paned.add(no_data_label)
            return
        
        # Left panel - Pie and Bar charts
        left_frame = tk.Frame(parent_paned, bg='white', relief='solid', bd=1)
        parent_paned.add(left_frame, weight=1)
        
        # Right panel - Line and other charts
        right_frame = tk.Frame(parent_paned, bg='white', relief='solid', bd=1)
        parent_paned.add(right_frame, weight=1)
        
        # Create left charts
        self.create_left_charts(left_frame, expense_data, monthly_data)
        
        # Create right charts
        self.create_right_charts(right_frame, monthly_data)
    
    def create_left_charts(self, parent, expense_data, monthly_data):
        
        fig = Figure(figsize=(8, 10), facecolor='white')
        
        # Colors
        colors_pie = ['#e74c3c', '#3498db', '#f39c12', '#2ecc71', '#9b59b6', '#1abc9c', '#e67e22', '#34495e']
        
        if expense_data:
            # Pie chart
            ax1 = fig.add_subplot(2, 1, 1)
            categories = [item[0] for item in expense_data[:8]]
            amounts = [abs(item[1]) for item in expense_data[:8]]
            
            wedges, texts, autotexts = ax1.pie(amounts, labels=categories, autopct='%1.1f%%',
                   startangle=90, colors=colors_pie[:len(categories)],
                   wedgeprops=dict(width=0.5, edgecolor='white', linewidth=2),
                   textprops={'fontsize': 10, 'weight': 'bold'})
            
            for autotext in autotexts:
                autotext.set_color('white')
            
            ax1.set_title('Kategoriye Göre Giderler', fontsize=14, weight='bold', pad=15)
            
            # Horizontal bar chart
            ax2 = fig.add_subplot(2, 1, 2)
            y_pos = np.arange(len(categories))
            colors_bar = plt.cm.Spectral(np.linspace(0.2, 0.8, len(categories)))
            
            bars = ax2.barh(y_pos, amounts, color=colors_bar, alpha=0.85, edgecolor='white', linewidth=2)
            
            ax2.set_yticks(y_pos)
            ax2.set_yticklabels(categories, fontsize=10)
            ax2.set_xlabel('Tutar (T)', fontsize=11, weight='bold')
            ax2.set_title('En Çok Harcanan Kategoriler', fontsize=14, weight='bold', pad=15)
            ax2.grid(True, alpha=0.3, axis='x', linestyle='--')
            ax2.set_facecolor('#f8f9fa')
            
            # Add value labels
            for i, (bar, value) in enumerate(zip(bars, amounts)):
                width = bar.get_width()
                ax2.text(width, bar.get_y() + bar.get_height()/2.,
                        f' ${value:,.0f}',
                        ha='left', va='center', fontsize=9, weight='bold')
        
        fig.tight_layout(pad=2.0)
        
        canvas = FigureCanvasTkAgg(fig, master=parent)
        canvas.draw()
        canvas_widget = canvas.get_tk_widget()
        canvas_widget.pack(fill=tk.BOTH, expand=True)
        
        # Add interactive tooltip with category data
        data_dict = {cat: amt for cat, amt in zip(categories, amounts)}
        self.add_category_tooltip(canvas, fig, data_dict)
    
    def create_right_charts(self, parent, monthly_data):

        fig = Figure(figsize=(8, 10), facecolor='white')
        
        months = [item[0] for item in monthly_data[-6:]]
        income = [item[1] for item in monthly_data[-6:]]
        expenses = [abs(item[2]) for item in monthly_data[-6:]]
        
        x = np.arange(len(months))
        
        # Line chart with area fill
        ax1 = fig.add_subplot(2, 1, 1)
        ax1.plot(x, income, marker='o', linewidth=3, markersize=8,
                color=self.colors['success'], label='Gelir', linestyle='-')
        ax1.fill_between(x, income, alpha=0.3, color=self.colors['success'])
        
        ax1.plot(x, expenses, marker='s', linewidth=3, markersize=8,
                color=self.colors['danger'], label='Gider', linestyle='-')
        ax1.fill_between(x, expenses, alpha=0.3, color=self.colors['danger'])
        
        ax1.set_xticks(x)
        ax1.set_xticklabels(months, rotation=45, ha='right', fontsize=9)
        ax1.set_title('Gelir ve Gider eğilimi', fontsize=14, weight='bold', pad=15)
        ax1.set_ylabel('Tatur (T)', fontsize=11, weight='bold')
        ax1.legend(loc='upper left', framealpha=0.9, fontsize=10)
        ax1.grid(True, alpha=0.3, linestyle='--')
        ax1.set_facecolor('#f8f9fa')
        
        # Net savings bar chart
        ax2 = fig.add_subplot(2, 1, 2)
        net = [i - e for i, e in zip(income, expenses)]
        colors_net = [self.colors['success'] if n >= 0 else self.colors['danger'] for n in net]
        
        bars = ax2.bar(x, net, color=colors_net, alpha=0.8, edgecolor='white', linewidth=2)
        ax2.axhline(y=0, color='black', linestyle='-', linewidth=1.5)
        
        # Add value labels
        for i, (bar, value) in enumerate(zip(bars, net)):
            height = bar.get_height()
            ax2.text(bar.get_x() + bar.get_width()/2., height,
                    f'${value:,.0f}',
                    ha='center', va='bottom' if value >= 0 else 'top',
                    fontsize=9, weight='bold')
        
        ax2.set_xticks(x)
        ax2.set_xticklabels(months, rotation=45, ha='right', fontsize=9)
        ax2.set_title('Net Tasarruf', fontsize=14, weight='bold', pad=15)
        ax2.set_ylabel('Tutar (T)', fontsize=11, weight='bold')
        ax2.grid(True, alpha=0.3, axis='y', linestyle='--')
        ax2.set_facecolor('#f8f9fa')
        
        fig.tight_layout(pad=2.0)
        
        canvas = FigureCanvasTkAgg(fig, master=parent)
        canvas.draw()
        canvas_widget = canvas.get_tk_widget()
        canvas_widget.pack(fill=tk.BOTH, expand=True)
        
        # Add interactive tooltip with date and amount data
        data_dict = {
            'months': months,
            'income': income,
            'expenses': expenses,
            'net': net
        }
        self.add_monthly_tooltip(canvas, fig, data_dict)
    
    def add_category_tooltip(self, canvas, fig, data_dict):

        tooltip = tk.Label(canvas.get_tk_widget(), text="",
                          bg='#2c3e50', fg='white',
                          font=('Segoe UI', 11, 'bold'),
                          padx=15, pady=10,
                          relief='solid', bd=2,
                          borderwidth=2,
                          highlightbackground='#3498db',
                          highlightthickness=2)
        tooltip.place_forget()

        def on_motion(event):
            if event.inaxes:
                y = event.ydata

                # Find closest category
                categories = list(data_dict.keys())
                if 0 <= y < len(categories):
                    idx = int(round(y))
                    if 0 <= idx < len(categories):
                        category = categories[idx]
                        amount = data_dict[category]

                        tooltip_text = f"🏷 {category}\n💰 ${amount:,.2f}"
                        tooltip.config(text=tooltip_text)

                        canvas_x = event.x + 20
                        canvas_y = event.y - 30
                        tooltip.place(x=canvas_x, y=canvas_y)
                        return

                tooltip.place_forget()
            else:
                tooltip.place_forget()

        def on_leave(event):
            tooltip.place_forget()

        canvas.mpl_connect('motion_notify_event', on_motion)
        canvas.mpl_connect('axes_leave_event', on_leave)

    def add_monthly_tooltip(self, canvas, fig, data_dict):
        """Add tooltip showing date and amounts"""

        tooltip = tk.Label(canvas.get_tk_widget(), text="",
                          bg='#2c3e50', fg='white',
                          font=('Segoe UI', 10, 'bold'),
                          padx=15, pady=10,
                          relief='solid', bd=2,
                          borderwidth=2,
                          highlightbackground='#3498db',
                          highlightthickness=2)
        tooltip.place_forget()

        months = data_dict['months']
        income = data_dict['income']
        expenses = data_dict['expenses']
        net = data_dict['net']

        def on_motion(event):
            if event.inaxes:
                x = event.xdata
                # Find closest month
                if x is not None and 0 <= x < len(months):
                    idx = int(round(x))
                    if 0 <= idx < len(months):
                        month = months[idx]
                        inc = income[idx]
                        exp = expenses[idx]
                        net_val = net[idx]

                        tooltip_text = f"📅 {month}\n"
                        tooltip_text += f"📈 Gelir: ${inc:,.2f}\n"
                        tooltip_text += f"📉 Gedirler: ${exp:,.2f}\n"
                        tooltip_text += f"💰 Net: ${net_val:,.2f}"

                        tooltip.config(text=tooltip_text)

                        canvas_x = event.x + 20
                        canvas_y = event.y - 60
                        tooltip.place(x=canvas_x, y=canvas_y)
                        return

                tooltip.place_forget()
            else:
                tooltip.place_forget()

        def on_leave(event):
            tooltip.place_forget()

        canvas.mpl_connect('motion_notify_event', on_motion)
        canvas.mpl_connect('axes_leave_event', on_leave)

    def create_expense_chart(self):
        # Get data
        expense_data = self.db.get_expenses_by_category()
        monthly_data = self.db.get_monthly_summary()

        if not expense_data and not monthly_data:
            return

        # Create figure with multiple subplots
        fig = Figure(figsize=(14, 10), facecolor='#fafafa')
        
        # Color schemes
        colors_pie = ['#e74c3c', '#3498db', '#f39c12', '#2ecc71', '#9b59b6', '#1abc9c', '#e67e22', '#34495e']
        
        if expense_data:
            # 1. Donut chart - Expenses by category
            ax1 = fig.add_subplot(2, 3, 1)
            categories = [item[0] for item in expense_data]
            amounts = [abs(item[1]) for item in expense_data]
            
            wedges, texts, autotexts = ax1.pie(amounts, labels=categories, autopct='%1.1f%%',
                   startangle=90, colors=colors_pie[:len(categories)],
                   wedgeprops=dict(width=0.5, edgecolor='white', linewidth=2),
                   textprops={'fontsize': 9, 'weight': 'bold'})
            
            for autotext in autotexts:
                autotext.set_color('white')
            
            ax1.set_title('Gider Dağılımı', fontsize=12, weight='bold', pad=15)
        
        if monthly_data and len(monthly_data) > 0:
            # 2. Line chart with area fill
            ax2 = fig.add_subplot(2, 3, 2)
            months = [item[0] for item in monthly_data[-6:]]
            income = [item[1] for item in monthly_data[-6:]]
            expenses = [abs(item[2]) for item in monthly_data[-6:]]
            
            x = np.arange(len(months))
            ax2.plot(x, income, marker='o', linewidth=2.5, markersize=7,
                    color='#2ecc71', label='Gelir', linestyle='-')
            ax2.fill_between(x, income, alpha=0.3, color='#2ecc71')
            
            ax2.plot(x, expenses, marker='s', linewidth=2.5, markersize=7,
                    color='#e74c3c', label='Giderler', linestyle='-')
            ax2.fill_between(x, expenses, alpha=0.3, color='#e74c3c')
            
            ax2.set_xticks(x)
            ax2.set_xticklabels(months, rotation=45, ha='right', fontsize=8)
            ax2.set_title('Gelir ve Gider Eğilimi', fontsize=12, weight='bold', pad=15)
            ax2.set_ylabel('Tutar (T)', fontsize=9)
            ax2.legend(loc='upper left', framealpha=0.9, fontsize=8)
            ax2.grid(True, alpha=0.3, linestyle='--')
            ax2.set_facecolor('#f8f9fa')
            
            # 3. Bar chart comparison
            ax3 = fig.add_subplot(2, 3, 3)
            width = 0.35
            ax3.bar(x - width/2, income, width, label='Gelir',
                   color='#2ecc71', alpha=0.8, edgecolor='white', linewidth=1.5)
            ax3.bar(x + width/2, expenses, width, label='Giderler',
                   color='#e74c3c', alpha=0.8, edgecolor='white', linewidth=1.5)
            
            ax3.set_xticks(x)
            ax3.set_xticklabels(months, rotation=45, ha='right', fontsize=8)
            ax3.set_title('Aylık Karşılaştırma', fontsize=12, weight='bold', pad=15)
            ax3.set_ylabel('Tutar (T)', fontsize=9)
            ax3.legend(framealpha=0.9, fontsize=8)
            ax3.grid(True, alpha=0.3, axis='y', linestyle='--')
            ax3.set_facecolor('#f8f9fa')
            
            # 4. Net savings
            ax4 = fig.add_subplot(2, 3, 4)
            net = [i - e for i, e in zip(income, expenses)]
            colors_net = ['#2ecc71' if n >= 0 else '#e74c3c' for n in net]
            
            bars = ax4.bar(x, net, color=colors_net, alpha=0.8, edgecolor='white', linewidth=1.5)
            ax4.axhline(y=0, color='black', linestyle='-', linewidth=1)
            
            # Add value labels
            for i, (bar, value) in enumerate(zip(bars, net)):
                height = bar.get_height()
                ax4.text(bar.get_x() + bar.get_width()/2., height,
                        f'${value:,.0f}',
                        ha='center', va='bottom' if value >= 0 else 'top',
                        fontsize=7, weight='bold')
            
            ax4.set_xticks(x)
            ax4.set_xticklabels(months, rotation=45, ha='right', fontsize=8)
            ax4.set_title('💰 Net Tasarruf', fontsize=12, weight='bold', pad=15)
            ax4.set_ylabel('Tutar (T)', fontsize=9)
            ax4.grid(True, alpha=0.3, axis='y', linestyle='--')
            ax4.set_facecolor('#f8f9fa')
            
            # 5. Horizontal bar - Categories
            if expense_data:
                ax5 = fig.add_subplot(2, 3, 5)
                categories_top = [item[0] for item in expense_data[:6]]
                amounts_top = [abs(item[1]) for item in expense_data[:6]]
                
                y_pos = np.arange(len(categories_top))
                colors_bar = plt.cm.Spectral(np.linspace(0.2, 0.8, len(categories_top)))
                
                bars = ax5.barh(y_pos, amounts_top, color=colors_bar, alpha=0.85, edgecolor='white', linewidth=1.5)
                
                ax5.set_yticks(y_pos)
                ax5.set_yticklabels(categories_top, fontsize=8)
                ax5.set_xlabel('Tutar (T)', fontsize=9)
                ax5.set_title('En Popüler Kategoriler', fontsize=12, weight='bold', pad=15)
                ax5.grid(True, alpha=0.3, axis='x', linestyle='--')
                ax5.set_facecolor('#f8f9fa')
                
                # Add value labels
                for i, (bar, value) in enumerate(zip(bars, amounts_top)):
                    width = bar.get_width()
                    ax5.text(width, bar.get_y() + bar.get_height()/2.,
                            f' ${value:,.0f}',
                            ha='left', va='center', fontsize=7, weight='bold')
            
            # 6. Cumulative analysis
            ax6 = fig.add_subplot(2, 3, 6)
            cumulative_income = np.cumsum(income)
            cumulative_expenses = np.cumsum(expenses)
            cumulative_net = cumulative_income - cumulative_expenses
            
            ax6.plot(x, cumulative_income, marker='o', linewidth=2,
                    color='#2ecc71', label='Kümülatif Gelir', linestyle='-')
            ax6.plot(x, cumulative_expenses, marker='s', linewidth=2,
                    color='#e74c3c', label='Kümülatif Giderler', linestyle='-')
            ax6.plot(x, cumulative_net, marker='^', linewidth=2.5,
                    color='#3498db', label='Net Pozisyon', linestyle='--')
            
            ax6.set_xticks(x)
            ax6.set_xticklabels(months, rotation=45, ha='right', fontsize=8)
            ax6.set_title('Kümülatif Analiz', fontsize=12, weight='bold', pad=15)
            ax6.set_ylabel('Tutar (T)', fontsize=9)
            ax6.legend(loc='upper left', framealpha=0.9, fontsize=8)
            ax6.grid(True, alpha=0.3, linestyle='--')
            ax6.set_facecolor('#f8f9fa')
        
        fig.tight_layout(pad=2.5)
        
        canvas = FigureCanvasTkAgg(fig, master=self.main_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    def show_add_transaction(self, trans_type):
        self.clear_frame()
        
        # Main container - full screen
        main_container = tk.Frame(self.main_frame, bg=self.colors['light'])
        main_container.pack(fill=tk.BOTH, expand=True)
        
        # Header with color based on transaction type
        header_color = self.colors['success'] if trans_type == "Income" else self.colors['danger']
        header_icon = "💰" if trans_type == "Gelir" else "💸"
        
        header = tk.Frame(main_container, bg=header_color, height=80)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        
        tk.Label(header, text=f"{header_icon} Add {trans_type}",
                font=('Segoe UI', 28, 'bold'),
                bg=header_color,
                fg='white').pack(side=tk.LEFT, padx=30, pady=20)
        
        # Cancel button in header
        cancel_btn = tk.Button(header, text="❌ İptal et",
                              command=self.show_dashboard,
                              bg='white',
                              fg=header_color,
                              font=('Segoe UI', 12, 'bold'),
                              relief='flat',
                              cursor='hand2',
                              padx=20,
                              pady=10)
        cancel_btn.pack(side=tk.RIGHT, padx=30, pady=20)
        
        def on_enter_cancel(e):
            cancel_btn['bg'] = '#ecf0f1'
        
        def on_leave_cancel(e):
            cancel_btn['bg'] = 'white'
        
        cancel_btn.bind('<Enter>', on_enter_cancel)
        cancel_btn.bind('<Leave>', on_leave_cancel)
        
        # Content container - full width with grid layout
        content = tk.Frame(main_container, bg='white')
        content.pack(fill=tk.BOTH, expand=True, padx=60, pady=40)
        
        # Configure grid for 2 columns
        content.grid_columnconfigure(0, weight=1)
        content.grid_columnconfigure(1, weight=1)
        
        # Left column - Amount and Category
        left_column = tk.Frame(content, bg='white')
        left_column.grid(row=0, column=0, sticky='nsew', padx=(0, 30))
        
        # Amount field
        amount_frame = tk.Frame(left_column, bg='white')
        amount_frame.pack(fill=tk.X, pady=15)
        
        tk.Label(amount_frame, text="💵 Tutar ",
                font=('Segoe UI', 14, 'bold'),
                bg='white',
                fg=self.colors['dark']).pack(anchor='w', pady=(0, 10))
        
        amount_container = tk.Frame(amount_frame, bg='#ecf0f1', relief='flat', bd=1)
        amount_container.pack(fill=tk.X)
        
        tk.Label(amount_container, text="T",
                font=('Segoe UI', 20, 'bold'),
                bg='#ecf0f1',
                fg=self.colors['primary']).pack(side=tk.LEFT, padx=(15, 10))
        
        amount_var = tk.DoubleVar()
        amount_entry = tk.Entry(amount_container, textvariable=amount_var,
                               font=('Segoe UI', 18),
                               bg='#ecf0f1',
                               fg=self.colors['dark'],
                               relief='flat',
                               bd=0)
        amount_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, pady=15, padx=(0, 15))
        amount_entry.focus()
        
        # Category field
        category_frame = tk.Frame(left_column, bg='white')
        category_frame.pack(fill=tk.X, pady=15)
        
        tk.Label(category_frame, text="🏷 Kategoriler",
                font=('Segoe UI', 14, 'bold'),
                bg='white',
                fg=self.colors['dark']).pack(anchor='w', pady=(0, 10))
        
        categories = self.db.get_categories()
        category_var = tk.StringVar()
        category_dropdown = ttk.Combobox(category_frame, textvariable=category_var,
                                       values=categories, state="readonly",
                                       font=('Segoe UI', 14),
                                       height=12)
        category_dropdown.pack(fill=tk.X, ipady=12)

        # Right column - Description and Date
        right_column = tk.Frame(content, bg='white')
        right_column.grid(row=0, column=1, sticky='nsew', padx=(30, 0))
        
        # Description field
        desc_frame = tk.Frame(right_column, bg='white')
        desc_frame.pack(fill=tk.X, pady=15)
        
        tk.Label(desc_frame, text="Açıklama (İsteğe Bağlı)",
                font=('Segoe UI', 14, 'bold'),
                bg='white',
                fg=self.colors['dark']).pack(anchor='w', pady=(0, 10))
        
        desc_container = tk.Frame(desc_frame, bg='#ecf0f1', relief='flat', bd=1)
        desc_container.pack(fill=tk.X)
        
        description_var = tk.StringVar()
        description_entry = tk.Entry(desc_container, textvariable=description_var,
                                     font=('Segoe UI', 14),
                                     bg='#ecf0f1',
                                     fg=self.colors['dark'],
                                     relief='flat',
                                     bd=0)
        description_entry.pack(fill=tk.X, pady=15, padx=15)
        
        # Date field with calendar
        date_frame = tk.Frame(right_column, bg='white')
        date_frame.pack(fill=tk.X, pady=15)
        
        tk.Label(date_frame, text="📅 Date",
                font=('Segoe UI', 14, 'bold'),
                bg='white',
                fg=self.colors['dark']).pack(anchor='w', pady=(0, 10))
        
        # Use DateEntry from tkcalendar for calendar widget
        date_entry = DateEntry(date_frame,
                              width=40,
                              background=self.colors['primary'],
                              foreground='white',
                              borderwidth=2,
                              font=('Segoe UI', 13),
                              date_pattern='yyyy-mm-dd',
                              state='readonly')
        date_entry.pack(fill=tk.X, ipady=12)
        
        # Save transaction function
        def save_transaction():
            try:
                amount = amount_var.get()
                category = category_var.get()
                description = description_var.get()
                date = date_entry.get_date().strftime("%Y-%m-%d")

                if not amount or amount <= 0:
                    messagebox.showerror("Hata", "Lütfen geçerli bir tutar girin!")
                    return
                
                if not category:
                    messagebox.showerror("Hata", "Lütfen bir kategori seçin!")
                    return

                self.db.add_transaction(
                    trans_type=trans_type,
                    amount=amount,
                    category=category,
                    description=description,
                    date=date
                )

                messagebox.showinfo("✅ Success", 
                                  f"{trans_type} of ${amount:,.2f} added successfully!\n\n"
                                  f"Category: {category}\n"
                                  f"Date: {date}")
                self.show_dashboard()
            except ValueError as e:
                messagebox.showerror("❌ Hata", "Please enter a valid amount!")
            except Exception as e:
                messagebox.showerror("❌ Hata", str(e))
        
        # Bottom section - Save button (full width)
        bottom_section = tk.Frame(content, bg='white')
        bottom_section.grid(row=1, column=0, columnspan=2, sticky='ew', pady=(40, 0))
        
        # Save button - large and centered
        save_btn = tk.Button(bottom_section, text=f"💾 Save {trans_type}",
                            command=save_transaction,
                            bg=header_color,
                            fg='white',
                            font=('Segoe UI', 16, 'bold'),
                            relief='flat',
                            cursor='hand2',
                            padx=60,
                            pady=20)
        save_btn.pack(expand=True)
        
        # Hover effects for save button
        def on_enter_save(e):
            save_btn['bg'] = self.colors['secondary']
        
        def on_leave_save(e):
            save_btn['bg'] = header_color
        
        save_btn.bind('<Enter>', on_enter_save)
        save_btn.bind('<Leave>', on_leave_save)


    def show_transactions(self):
        self.clear_frame()

        ttk.Label(self.main_frame, text="All Transactions",
                  font=("Arial", 16, "bold")).pack(pady=10)

        # Add filter options
        filter_frame = ttk.Frame(self.main_frame)
        filter_frame.pack(fill=tk.X, pady=10)

        # Type filter
        ttk.Label(filter_frame, text="Type:").pack(side=tk.LEFT, padx=5)
        type_var = tk.StringVar()
        type_dropdown = ttk.Combobox(filter_frame, textvariable=type_var,
                                   values=["All", "Income", "Expense"],
                                   state="readonly", width=10)
        type_dropdown.set("All")
        type_dropdown.pack(side=tk.LEFT, padx=5)

        # Category filter
        ttk.Label(filter_frame, text="Category:").pack(side=tk.LEFT, padx=5)
        category_var = tk.StringVar()
        categories = ["All"] + self.db.get_categories()
        category_dropdown = ttk.Combobox(filter_frame, textvariable=category_var,
                                       values=categories, state="readonly", width=15)
        category_dropdown.set("All")
        category_dropdown.pack(side=tk.LEFT, padx=5)

        # Date range filter
        ttk.Label(filter_frame, text="From:").pack(side=tk.LEFT, padx=5)
        from_date_var = tk.StringVar()
        ttk.Entry(filter_frame, textvariable=from_date_var, width=10).pack(side=tk.LEFT, padx=5)

        ttk.Label(filter_frame, text="To:").pack(side=tk.LEFT, padx=5)
        to_date_var = tk.StringVar()
        ttk.Entry(filter_frame, textvariable=to_date_var, width=10).pack(side=tk.LEFT, padx=5)

        # Create treeview for transactions
        columns = ("ID", "Tarih", "Tür", "Kategori", "Tutar", "Açıklama")
        tree = ttk.Treeview(self.main_frame, columns=columns, show="headings", height=20)

        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=120)

        tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Function to load and filter transactions
        def load_transactions():
            # Clear existing items
            for item in tree.get_children():
                tree.delete(item)
            
            # Get all transactions
            all_transactions = self.db.get_all_transactions()
            
            # Apply filters
            selected_type = type_var.get()
            selected_category = category_var.get()
            from_date = from_date_var.get().strip()
            to_date = to_date_var.get().strip()
            
            for trans in all_transactions:
                # trans = (id, date, type, category, amount, description)
                trans_id, trans_date, trans_type, trans_category, trans_amount, trans_desc = trans
                
                # Filter by type
                if selected_type != "Hepsi" and trans_type != selected_type:
                    continue
                
                # Filter by category
                if selected_category != "Hepsi" and trans_category != selected_category:
                    continue
                
                # Filter by date range
                if from_date:
                    try:
                        if trans_date < from_date:
                            continue
                    except:
                        pass
                
                if to_date:
                    try:
                        if trans_date > to_date:
                            continue
                    except:
                        pass
                
                # Add to tree
                tree.insert("", "end", values=trans)

        # Search button
        def search_transactions():
            load_transactions()

        ttk.Button(filter_frame, text="Ara", command=search_transactions).pack(side=tk.LEFT, padx=10)
        
        # Reset button
        def reset_filters():
            type_var.set("All")
            category_var.set("All")
            from_date_var.set("")
            to_date_var.set("")
            load_transactions()
        
        ttk.Button(filter_frame, text="Sıfırla", command=reset_filters).pack(side=tk.LEFT, padx=5)

        # Load all transactions initially
        load_transactions()

        # Back button
        ttk.Button(self.main_frame, text="Panoya Geri Dön",
                   command=self.show_dashboard).pack(pady=10)


    def show_report(self, period):
        self.clear_frame()

        ttk.Label(self.main_frame, text=f"{period} Repor",
                  font=("Arial", 16, "bold")).pack(pady=10)

        # Generate report data based on period
        if period == "Daily":
            data = self.db.get_daily_summary()
            x_label = "Gün"
        elif period == "Monthly":
            data = self.db.get_monthly_summary()
            x_label = "Ay"
        else:  # Yearly
            data = self.db.get_yearly_summary()
            x_label = "Yıl"

        if not data:
            ttk.Label(self.main_frame, text="Veri mevcut değil").pack(pady=20)
            ttk.Button(self.main_frame, text="Geri",
                       command=self.show_dashboard).pack(pady=10)
            return

        # Create figure
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 8))

        # Plot income vs expenses
        dates = [item[0] for item in data]
        income = [item[1] for item in data]
        expenses = [abs(item[2]) for item in data]  # Make expenses positive for the chart

        ax1.bar(dates, income, width=0.4, label='Gelir', color='green')
        ax1.bar([str(d) for d in dates], expenses, width=0.4, label='Gedir', color='red',
                bottom=income)
        ax1.set_ylabel('Tutar (T)')
        ax1.set_title(f'{period} Gelir vs Gedir')
        ax1.legend()

        # Plot net income
        net = [i - e for i, e in zip(income, expenses)]
        ax2.bar(dates, net, color='blue' if net[-1] >= 0 else 'red')
        ax2.axhline(0, color='black', linewidth=0.5)
        ax2.set_xlabel(x_label)
        ax2.set_ylabel('Net Gelir (T)')
        ax2.set_title(f'{period}Net Gelir')

        plt.tight_layout()

        # Create canvas and add to tkinter window
        canvas = FigureCanvasTkAgg(fig, master=self.main_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Back button
        ttk.Button(self.main_frame, text="Panoya geri dön",
                   command=self.show_dashboard).pack(pady=10)

    def show_statistics(self):
        self.clear_frame()
        
        ttk.Label(self.main_frame, text="📊 Analiz",
                 font=("Arial", 20, "bold")).pack(pady=15)
        
        all_transactions = self.db.get_all_transactions()
        
        if not all_transactions:
            ttk.Label(self.main_frame, text="Hem gelir hem de gider verisi gerekli",
                     font=("Arial", 12)).pack(pady=50)
            ttk.Button(self.main_frame, text="Geri", command=self.show_dashboard).pack(pady=10)
            return
        
        income_amounts = [t[4] for t in all_transactions if t[2] == 'Income']
        expense_amounts = [t[4] for t in all_transactions if t[2] == 'Expense']
        
        if not income_amounts or not expense_amounts:
            ttk.Label(self.main_frame, text="Hem gelir hem de gider verisi gerekli",
                     font=("Arial", 12)).pack(pady=50)
            ttk.Button(self.main_frame, text="Geri", command=self.show_dashboard).pack(pady=10)
            return
        
        fig = Figure(figsize=(14, 10), facecolor='#fafafa')
        
        ax1 = fig.add_subplot(2, 3, 1)
        ax1.hist(income_amounts, bins=20, color='#2ecc71', alpha=0.7, edgecolor='black')
        ax1.set_title('📈 Gelir Dağılımı', fontsize=12, weight='bold')
        ax1.set_xlabel('Tutar (T)')
        ax1.set_ylabel('Sıklık')
        ax1.grid(True, alpha=0.3)
        mean_income = np.mean(income_amounts)
        ax1.axvline(mean_income, color='red', linestyle='--', linewidth=2, label=f'Ortalama: ${mean_income:.2f}')
        ax1.legend()
        
        ax2 = fig.add_subplot(2, 3, 2)
        ax2.hist(expense_amounts, bins=20, color='#e74c3c', alpha=0.7, edgecolor='black')
        ax2.set_title('📉 Gider Dağılımı', fontsize=12, weight='bold')
        ax2.set_xlabel('Tutar (T)')
        ax2.set_ylabel('Sıklık')
        ax2.grid(True, alpha=0.3)
        mean_expense = np.mean(expense_amounts)
        ax2.axvline(mean_expense, color='darkred', linestyle='--', linewidth=2, label=f'Ortalama: ${mean_expense:.2f}')
        ax2.legend()
        
        ax3 = fig.add_subplot(2, 3, 3)
        bp = ax3.boxplot([income_amounts, expense_amounts], labels=['Gelir', 'Gider'], patch_artist=True)
        bp['boxes'][0].set_facecolor('#2ecc71')
        bp['boxes'][1].set_facecolor('#e74c3c')
        ax3.set_title('📦 Kutu Grafiği', fontsize=12, weight='bold')
        ax3.set_ylabel('Tutar (T)')
        ax3.grid(True, alpha=0.3, axis='y')
        
        fig.tight_layout(pad=2.5)
        
        canvas=FigureCanvasTkAgg(fig, master=self.main_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        ttk.Button(self.main_frame, text="Geri", command=self.show_dashboard).pack(pady=10)
    
    def show_3d_analysis(self):
        self.clear_frame()
        
        ttk.Label(self.main_frame, text="🎯 3D Analiz",
                 font=("Arial", 20, "bold")).pack(pady=15)
        
        monthly_data = self.db.get_monthly_summary()
        
        if not monthly_data or len(monthly_data) < 3:
            ttk.Label(self.main_frame, text="3D boyutlu görselleştirme için yeterli veri yok",
                     font=("Arial", 12)).pack(pady=50)
            ttk.Button(self.main_frame, text="Geri", command=self.show_dashboard).pack(pady=10)
            return
        fig = Figure(figsize=(14, 10), facecolor='#fafafa')
        
        ax1 = fig.add_subplot(1, 2, 1, projection='3d')
        
        months = [item[0] for item in monthly_data[-12:]]
        income = [item[1] for item in monthly_data[-12:]]
        expenses = [abs(item[2]) for item in monthly_data[-12:]]
        
        x_pos = np.arange(len(months))
        y_pos = np.array([0, 1])
        
        xpos, ypos = np.meshgrid(x_pos, y_pos)
        xpos = xpos.flatten()
        ypos = ypos.flatten()
        zpos = np.zeros_like(xpos)
        
        dx = 0.4 * np.ones_like(zpos)
        dy = 0.4 * np.ones_like(zpos)
        
        dz = []
        colors = []
        for i in range(len(months)):
            dz.extend([income[i], expenses[i]])
            colors.extend(['#2ecc71', '#e74c3c'])
        
        ax1.bar3d(xpos, ypos, zpos, dx, dy, dz, color=colors, alpha=0.8)
        ax1.set_xlabel('Ay')
        ax1.set_ylabel('Tür')
        ax1.set_zlabel('Tutar (T)')
        ax1.set_title('3D Boyutlu Gelir ve Giderler', fontsize=12, weight='bold')
        ax1.set_yticks([0, 1])
        ax1.set_yticklabels(['Gelir', 'Giderler'])
        
        ax2 = fig.add_subplot(1, 2, 2, projection='3d')
        
        X = np.arange(len(months))
        Y = np.array([0, 1, 2])
        X, Y = np.meshgrid(X, Y)
        
        Z = np.array([income, expenses, [i-e for i, e in zip(income, expenses)]])
        
        surf = ax2.plot_surface(X, Y, Z, cmap='viridis', alpha=0.8)
        ax2.set_xlabel('Ay')
        ax2.set_ylabel('Ölçüt')
        ax2.set_zlabel('Tutar (T)')
        ax2.set_title('🌊 Finansal Yüzey', fontsize=12, weight='bold')
        ax2.set_yticks([0, 1, 2])
        ax2.set_yticklabels(['Gelir', 'Giderler', 'Net'])
        
        fig.tight_layout(pad=2.0)
        
        canvas = FigureCanvasTkAgg(fig, master=self.main_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        ttk.Button(self.main_frame, text="Geri", command=self.show_dashboard).pack(pady=10)

    def manage_categories(self):
        self.clear_frame()
        
        # Main container - full screen
        main_container = tk.Frame(self.main_frame, bg=self.colors['light'])
        main_container.pack(fill=tk.BOTH, expand=True)
        
        # Header - full width
        header = tk.Frame(main_container, bg=self.colors['secondary'], height=80)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        
        tk.Label(header, text="🏷 Kategorileri Yönet",
                font=('Segoe UI', 28, 'bold'),
                bg=self.colors['secondary'],
                fg='white').pack(side=tk.LEFT, padx=30, pady=20)
        
        # Back button in header
        back_btn = tk.Button(header, text="Panoya geri dön",
                            command=self.show_dashboard,
                            bg='white',
                            fg=self.colors['secondary'],
                            font=('Segoe UI', 12, 'bold'),
                            relief='flat',
                            cursor='hand2',
                            padx=20,
                            pady=10)
        back_btn.pack(side=tk.RIGHT, padx=30, pady=20)
        
        def on_enter_back(e):
            back_btn['bg'] = '#ecf0f1'
        
        def on_leave_back(e):
            back_btn['bg'] = 'white'
        
        back_btn.bind('<Enter>', on_enter_back)
        back_btn.bind('<Leave>', on_leave_back)
        
        # Content container - full width
        content = tk.Frame(main_container, bg='white')
        content.pack(fill=tk.BOTH, expand=True, padx=40, pady=30)
        
        # Add Category Section
        add_section = tk.Frame(content, bg='white')
        add_section.pack(fill=tk.X, pady=(0, 30))
        
        tk.Label(add_section, text="Yeni Kategori Ekle",
                font=('Segoe UI', 16, 'bold'),
                bg='white',
                fg=self.colors['primary']).pack(anchor='w', pady=(0, 15))
        
        # Input container
        input_container = tk.Frame(add_section, bg='white')
        input_container.pack(fill=tk.X)
        
        # Category name input
        input_frame = tk.Frame(input_container, bg='#ecf0f1', relief='flat', bd=1)
        input_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        tk.Label(input_frame, text="🏷",
                font=('Segoe UI', 16),
                bg='#ecf0f1',
                fg=self.colors['primary']).pack(side=tk.LEFT, padx=(10, 5))
        
        new_category_var = tk.StringVar()
        category_entry = tk.Entry(input_frame, textvariable=new_category_var,
                                  font=('Segoe UI', 14),
                                  bg='#ecf0f1',
                                  fg=self.colors['dark'],
                                  relief='flat',
                                  bd=0)
        category_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, pady=12, padx=(0, 10))
        category_entry.insert(0, "Kategori adını girin...")
        
        # Placeholder behavior
        def on_entry_click(event):
            if category_entry.get() == "Kategori adını girin...":
                category_entry.delete(0, tk.END)
                category_entry.config(fg=self.colors['dark'])
        
        def on_focusout(event):
            if category_entry.get() == "":
                category_entry.insert(0, "Kategori adını girin...")
                category_entry.config(fg='gray')
        
        category_entry.bind('<FocusIn>', on_entry_click)
        category_entry.bind('<FocusOut>', on_focusout)
        category_entry.config(fg='gray')
        
        # Add button
        def add_category():
            category = new_category_var.get().strip()
            if category and category != "Kategori adını girin...":
                try:
                    self.db.add_category(category)
                    messagebox.showinfo("✅ Başarılı", f"{category}' kategorisi başarıyla eklendi!")
                    new_category_var.set("")
                    category_entry.delete(0, tk.END)
                    category_entry.insert(0, "Kategori adını girin...")
                    category_entry.config(fg='gray')
                    refresh_categories()
                except sqlite3.IntegrityError:
                    messagebox.showerror("❌ Hata", "Kategori zaten mevcut")
            else:
                messagebox.showerror("❌ Hata", "Lütfen bir kategori adı girin")
        
        add_btn = tk.Button(input_container, text="➕ Ekle",
                           command=add_category,
                           bg=self.colors['success'],
                           fg='white',
                           font=('Segoe UI', 12, 'bold'),
                           relief='flat',
                           cursor='hand2',
                           padx=30,
                           pady=12)
        add_btn.pack(side=tk.LEFT)
        
        # Hover effect for add button
        def on_enter_add(e):
            add_btn['bg'] = '#27ae60'
        
        def on_leave_add(e):
            add_btn['bg'] = self.colors['success']
        
        add_btn.bind('<Enter>', on_enter_add)
        add_btn.bind('<Leave>', on_leave_add)
        
        # Separator
        separator = tk.Frame(content, bg='#bdc3c7', height=2)
        separator.pack(fill=tk.X, pady=20)
        
        # Categories List Section
        list_section = tk.Frame(content, bg='white')
        list_section.pack(fill=tk.BOTH, expand=True)
        
        tk.Label(list_section, text="Mevcut Kategoriler",
                font=('Segoe UI', 16, 'bold'),
                bg='white',
                fg=self.colors['primary']).pack(anchor='w', pady=(0, 15))
        
        # Scrollable categories container - full height
        categories_canvas = tk.Canvas(list_section, bg='white', highlightthickness=0)
        categories_scrollbar = ttk.Scrollbar(list_section, orient="vertical", command=categories_canvas.yview)
        categories_frame = tk.Frame(categories_canvas, bg='white')


        categories_frame.bind("<Configure>", lambda e: categories_canvas.configure(scrollregion=categories_canvas.bbox("all")))
        categories_canvas.create_window((0, 0), window=categories_frame, anchor="nw")
        categories_canvas.configure(yscrollcommand=categories_scrollbar.set)
        
        categories_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        categories_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        def refresh_categories():
            for widget in categories_frame.winfo_children():
                widget.destroy()
            
            categories = self.db.get_categories()
            
            if not categories:
                no_cat_label = tk.Label(categories_frame, text="Henüz kategori yok. Yukarıdan bir tane ekleyin!",
                                       font=('Segoe UI', 12),
                                       bg='white',
                                       fg='gray')
                no_cat_label.pack(pady=50)
                return
            
            for i, category in enumerate(categories):
                # Category card
                cat_card = tk.Frame(categories_frame, bg='#f8f9fa', relief='solid', bd=1)
                cat_card.pack(fill=tk.X, pady=5, padx=5)
                
                # Left side - icon and name
                left_frame = tk.Frame(cat_card, bg='#f8f9fa')
                left_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=15, pady=12)
                
                # Category icon
                category_icons = {
                    'Maaş': '💰',  # Salary
                    'Serbest Çalışma': '💼',  # Freelance
                    'Yatırımlar': '📈',  # Investments
                    'Konut': '🏠',  # Housing
                    'Yiyecek': '🍔',  # Food
                    'Ulaşım': '🚗',  # Transportation
                    'Fatura': '💡',  # Utilities (Literally "Bill")
                    'Eğlence': '🎬',  # Entertainment
                    'Sağlık': '⚕️',  # Health
                    'Alışveriş': '🛍',  # Shopping
                    'Diğer': '📦'  # Others
                }
                icon = category_icons.get(category, '🏷')
                
                tk.Label(left_frame, text=icon,
                        font=('Segoe UI', 20),
                        bg='#f8f9fa').pack(side=tk.LEFT, padx=(0, 10))
                
                tk.Label(left_frame, text=category,
                        font=('Segoe UI', 13, 'bold'),
                        bg='#f8f9fa',
                        fg=self.colors['dark']).pack(side=tk.LEFT)
                
                # Right side - status and delete button
                right_frame = tk.Frame(cat_card, bg='#f8f9fa')
                right_frame.pack(side=tk.RIGHT, padx=15, pady=12)
                
                in_use = self.db.is_category_in_use(category)
                
                if in_use:
                    # In use badge
                    badge = tk.Label(right_frame, text="Kullanımda",
                                    font=('Segoe UI', 9, 'bold'),
                                    bg=self.colors['success'],
                                    fg='white',
                                    padx=10,
                                    pady=5)
                    badge.pack(side=tk.LEFT, padx=5)
                else:
                    # Delete button
                    def delete_category(cat=category):
                        if messagebox.askyesno("⚠️ Silme Onayı",
                                             f"'{cat}' kategorisini silmek istediğinizden emin misiniz?\n\nBu eylem geri alınamaz."):
                            self.db.delete_category(cat)
                            messagebox.showinfo("✅ Başarılı", f" '{cat}' kategorisi başarıyla silindi!")
                            refresh_categories()
                    
                    delete_btn = tk.Button(right_frame, text="🗑 Sil",
                                          command=delete_category,
                                          bg=self.colors['danger'],
                                          fg='white',
                                          font=('Segoe UI', 10, 'bold'),
                                          relief='flat',
                                          cursor='hand2',
                                          padx=15, pady=5)
                    delete_btn.pack(side=tk.LEFT, padx=5)
                    
                    # Hover effect
                    def on_enter_del(e, btn=delete_btn):
                        btn['bg'] = '#c0392b'
                    
                    def on_leave_del(e, btn=delete_btn):
                        btn['bg'] = self.colors['danger']
                    
                    delete_btn.bind('<Enter>', on_enter_del)
                    delete_btn.bind('<Leave>', on_leave_del)
        
        refresh_categories()



class Database:
    def __init__(self, db_file="finance.db"):
        self.db_file = db_file
        self.create_tables()
        self.initialize_default_categories()

    def get_connection(self):
        return sqlite3.connect(self.db_file)

    def create_tables(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()

            # Categories table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS categories (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT UNIQUE NOT NULL
                )
            ''')

            # Transactions table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS transactions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    date DATE NOT NULL,
                    type TEXT NOT NULL,  -- 'Income' or 'Expense'
                    category_id INTEGER NOT NULL,
                    amount REAL NOT NULL,
                    description TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (category_id) REFERENCES categories (id)
                )
            ''')

            conn.commit()

    def initialize_default_categories(self):
        default_categories = [
            "Maaş", "Serbest Çalışma", "Yatırımlar",  # Gelir Kategorileri (Income categories)
            "Konut", "Yiyecek", "Ulaşım",  # Gider Kategorileri (Expense categories)
            "Fatura", "Eğlence", "Sağlık", "Alışveriş", "Diğer"
        ]

        with self.get_connection() as conn:
            cursor = conn.cursor()
            for category in default_categories:
                try:
                    cursor.execute("INSERT INTO categories (name) VALUES (?)", (category,))
                except sqlite3.IntegrityError:
                    pass  # Category already exists
            conn.commit()

    def add_category(self, name):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("INSERT INTO categories (name) VALUES (?)", (name,))
            conn.commit()

    def delete_category(self, name):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM categories WHERE name = ?", (name,))
            conn.commit()

    def get_categories(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM categories ORDER BY name")
            return [row[0] for row in cursor.fetchall()]

    def get_category_id(self, name):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM categories WHERE name = ?", (name,))
            result = cursor.fetchone()
            return result[0] if result else None

    def is_category_in_use(self, name):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT COUNT(*) FROM transactions 
                WHERE category_id = (SELECT id FROM categories WHERE name = ?)
            """, (name,))
            return cursor.fetchone()[0] > 0

    def add_transaction(self, trans_type, amount, category, description="", date=None):
        if not date:
            date = datetime.now().strftime("%Y-%m-%d")

        category_id = self.get_category_id(category)
        if not category_id:
            raise ValueError(f"'{category}' kategorisi mevcut değil")
        
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO transactions (date, type, category_id, amount, description)
                VALUES (?, ?, ?, ?, ?)
            """, (date, trans_type, category_id, amount, description))
            conn.commit()

    def get_balance(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()

            # Get total income
            cursor.execute("SELECT COALESCE(SUM(amount), 0) FROM transactions WHERE type = 'Income'")
            total_income = cursor.fetchone()[0] or 0

            # Get total expenses
            cursor.execute("SELECT COALESCE(SUM(amount), 0) FROM transactions WHERE type = 'Expense'")
            total_expenses = cursor.fetchone()[0] or 0

            return total_income - total_expenses

    def get_recent_transactions(self, limit=10):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT t.date, t.type, c.name, t.amount, t.description
                FROM transactions t
                JOIN categories c ON t.category_id = c.id
                ORDER BY t.date DESC, t.created_at DESC
                LIMIT ?
            """, (limit,))
            return cursor.fetchall()

    def get_category_summary(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT 
                    c.name as category,
                    t.type,
                    SUM(t.amount) as total_amount
                FROM transactions t
                JOIN categories c ON t.category_id = c.id
                GROUP BY c.name, t.type
                ORDER BY t.type, total_amount DESC
            """)
            columns = [desc[0] for desc in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]

    def get_all_transactions(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                           SELECT t.id, t.date, t.type, c.name, t.amount, t.description
                           FROM transactions t
                                    JOIN categories c ON t.category_id = c.id
                           ORDER BY t.date DESC, t.created_at DESC
                           """)
            return cursor.fetchall()

    def get_expenses_by_category(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT c.name, SUM(t.amount) as total
                FROM transactions t
                JOIN categories c ON t.category_id = c.id
                WHERE t.type = 'Expense'
                GROUP BY c.name
                HAVING total > 0
                ORDER BY total DESC
            """)
            return cursor.fetchall()
    def get_daily_summary(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT 
                    date,
                    COALESCE(SUM(CASE WHEN type = 'Income' THEN amount ELSE 0 END), 0) as income,
                    COALESCE(SUM(CASE WHEN type = 'Expense' THEN amount ELSE 0 END), 0) as expenses
                FROM transactions
                GROUP BY date
                ORDER BY date
            """)
            return cursor.fetchall()

    def get_monthly_summary(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT 
                    strftime('%Y-%m', date) as month,
                    COALESCE(SUM(CASE WHEN type = 'Income' THEN amount ELSE 0 END), 0) as income,
                    COALESCE(SUM(CASE WHEN type = 'Expense' THEN amount ELSE 0 END), 0) as expenses
                FROM transactions
                GROUP BY month
                ORDER BY month
            """)
            return cursor.fetchall()

    def get_yearly_summary(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT 
                    strftime('%Y', date) as year,
                    COALESCE(SUM(CASE WHEN type = 'Income' THEN amount ELSE 0 END), 0) as income,
                    COALESCE(SUM(CASE WHEN type = 'Expense' THEN amount ELSE 0 END), 0) as expenses
                FROM transactions
                GROUP BY year
                ORDER BY year
            """)
            return cursor.fetchall()
    
    def get_total_income(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT COALESCE(SUM(amount), 0) FROM transactions WHERE type = 'Income'")
            return cursor.fetchone()[0] or 0
    
    def get_total_expenses(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT COALESCE(SUM(amount), 0) FROM transactions WHERE type = 'Expense'")
            return cursor.fetchone()[0] or 0
    
    def get_monthly_expenses(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            current_month = datetime.now().strftime("%Y-%m")
            cursor.execute("""
                SELECT COALESCE(SUM(amount), 0) 
                FROM transactions 
                WHERE type = 'Expense' AND strftime('%Y-%m', date) = ?
            """, (current_month,))
            return cursor.fetchone()[0] or 0
    
    def get_monthly_income(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            current_month = datetime.now().strftime("%Y-%m")
            cursor.execute("""
                SELECT COALESCE(SUM(amount), 0) 
                FROM transactions 
                WHERE type = 'Income' AND strftime('%Y-%m', date) = ?
            """, (current_month,))
            return cursor.fetchone()[0] or 0


if __name__ == "__main__":
    root = tk.Tk()
    app = FinanceTracker(root)
    root.mainloop()
